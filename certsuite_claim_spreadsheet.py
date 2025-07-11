#!/usr/bin/env python3
import argparse
import json, os, sys
import re, fileinput
import subprocess
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.worksheet import Worksheet
from typing import Dict, List, Any, Tuple

def trim_empty_lines(text: str) -> str:
    """Remove empty lines from text."""
    lines = text.split('\n')
    trimmed_lines = [line for line in lines if line]
    return '\n'.join(trimmed_lines)

def replace_text_in_file(input_file: str, search_text: str, replace_text: str, output_file: str) -> None:
    """Replace text in a file and save to output file."""
    with open(input_file, 'r') as f:
        # Read the entire contents of the file
        filedata = f.read()

    # Replace the search text with the replace text
    newdata = filedata.replace(search_text, replace_text)

    with open(output_file, 'w') as f:
        # Write the modified data to the output file
        f.write(newdata)

def extract_test_results(data: Dict[str, Any]) -> Tuple[List[Dict[str, Any]], int, int, int, int]:
    """Extract and process test results from the claim data."""
    try:
        results = []
        claim_results = data.get('claim', {}).get('results', {})
        
        if not claim_results:
            raise ValueError("No test results found in claim data")
            
        for key, test in claim_results.items():
            try:
                result = {
                    'Test_Id': test.get('testID', {}).get('id', ''),
                    'Test_Text': test.get('catalogInfo', {}).get('description', ''),
                    'State': test.get('state', ''),
                    'Category_Classification': ', '.join(f"{k}: {v}" for k, v in test.get('categoryClassification', {}).items()),
                    'Exception_Process': test.get('catalogInfo', {}).get('exceptionProcess', ''),
                    'Remediation': test.get('catalogInfo', {}).get('remediation', '')
                }
                
                # Filter out lines containing "INFO" from capturedTestOutput
                captured_output = test.get('capturedTestOutput', '')
                if captured_output:
                    output_lines = captured_output.strip().split('\n')
                    filtered_output = [line for line in output_lines if "INFO" not in line]
                    filtered_output_str = '\n'.join(filtered_output)
                else:
                    filtered_output_str = ''
                
                # Exclude 'Capture_Output' if state is 'passed'
                if test.get('state') != 'passed':
                    result['Capture_Output'] = filtered_output_str
                    result['Best_Practice_Link'] = test.get('catalogInfo', {}).get('bestPracticeReference', '')

                results.append(result)
            except Exception as e:
                print(f"Warning: Error processing test {key}: {e}")
                continue

        # Sort tests by state (failed, error, skipped, passed) and then alphabetically
        failed_tests = sorted([r for r in results if r['State'] == 'failed'], key=lambda r: r['Test_Id'])
        error_tests = sorted([r for r in results if r['State'] == 'error'], key=lambda r: r['Test_Id'])
        skipped_tests = sorted([r for r in results if r['State'] == 'skipped'], key=lambda r: r['Test_Id'])
        passed_tests = sorted([r for r in results if r['State'] == 'passed'], key=lambda r: r['Test_Id'])

        # Calculate totals
        total_failed = len(failed_tests)
        total_error = len(error_tests)
        total_skipped = len(skipped_tests)
        total_passed = len(passed_tests)

        # Combine sorted tests into a single list (failed, error, skipped, passed)
        sorted_tests = failed_tests + error_tests + skipped_tests + passed_tests
        
        return sorted_tests, total_failed, total_error, total_skipped, total_passed
    
    except Exception as e:
        raise ValueError(f"Error extracting test results: {e}")

def create_workbook_and_worksheet(output_file: str) -> Tuple[openpyxl.Workbook, Worksheet]:
    """Create Excel workbook and worksheet with proper naming."""
    wb = openpyxl.Workbook()
    ws = wb.active
    if ws is None:
        ws = wb.create_sheet()
    
    sheetname = os.path.basename(output_file)
    ws.title = sheetname.strip()
    
    return wb, ws

def add_test_results_to_worksheet(ws: Worksheet, sorted_tests: List[Dict[str, Any]]) -> None:
    """Add test results to the worksheet."""
    # Define column headers
    headers = ['Test_Id', 'Test_Text', 'State', 'Capture_Output', 'Category_Classification', 'Exception_Process', 'Remediation', 'Best_Practice_Link']
    
    # Add headers to worksheet
    ws.append(headers)
    
    # Add test results to worksheet
    for test in sorted_tests:
        row_data = []
        for header in headers:
            # Get value from test dictionary, use empty string if key doesn't exist
            value = test.get(header, '')
            row_data.append(value)
        ws.append(row_data)

def apply_basic_styling(ws: Worksheet) -> Dict[str, Any]:
    """Apply basic styling to the worksheet and return style objects."""
    # Define font and fill styles
    bold_font = Font(bold=True)
    red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    dark_red_fill = PatternFill(start_color='8B0000', end_color='8B0000', fill_type='solid')  # Dark red for errors
    orange_fill = PatternFill(start_color='FFE599', end_color='FFE599', fill_type='solid')
    blue_fill = PatternFill(start_color='AED6F1', end_color='AED6F1', fill_type='solid')
    green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
    yellow_fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
    light_green_fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')

    arial_font = Font(name='Arial')
    header_font = Font(bold=True)
    header_border = Border(left=Side(border_style='medium', color='000000'),
                    right=Side(border_style='medium', color='000000'),
                    top=Side(border_style='medium', color='000000'),
                    bottom=Side(border_style='medium', color='000000'))

    # Set cell styles based on test state
    for row in ws.iter_rows(min_row=1):
        for cell in row:
            cell.font = arial_font
            if cell.row == 1:
                cell.font = header_font
                cell.fill = blue_fill
                cell.border = header_border
            elif cell.column == 3:
                if cell.value == 'failed':
                    cell.fill = red_fill
                elif cell.value == 'error':
                    cell.fill = dark_red_fill
                elif cell.value == 'skipped':
                    cell.fill = orange_fill
                elif cell.value == 'passed':
                    cell.fill = green_fill
    
    return {
        'bold_font': bold_font,
        'red_fill': red_fill,
        'dark_red_fill': dark_red_fill,
        'orange_fill': orange_fill,
        'blue_fill': blue_fill,
        'green_fill': green_fill,
        'yellow_fill': yellow_fill,
        'light_green_fill': light_green_fill,
        'arial_font': arial_font,
        'header_font': header_font,
        'header_border': header_border
    }

def set_column_formatting(ws: Worksheet) -> None:
    """Set column width and alignment."""
    # Set column width and alignment
    for column in ws.columns:
        max_length = 0
        if column[0].column is not None:
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                     if len(str(cell.value)) > max_length:
                         max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width
            if column[0].column == 3:
                for cell in column:
                    cell.alignment = Alignment(horizontal='center')
            else:
                for cell in column:
                    cell.alignment = Alignment(wrapText=True)

def add_summary_section(ws: Worksheet, sorted_tests: List[Dict[str, Any]], 
                       failed_tests: int, error_tests: int, skipped_tests: int, passed_tests: int, 
                       dci_jobid: str, styles: Dict[str, Any]) -> None:
    """Add summary section at the top of the worksheet."""
    # Add summary section
    ws.insert_rows(1, amount=9)  # Insert 9 empty rows at the top (one more for error)
    ws['A1'] = 'Summary'
    ws['A2'] = 'Total'
    ws['A3'] = 'Failed'
    ws['A4'] = 'Error'
    ws['A5'] = 'Skipped'
    ws['A6'] = 'Passed'
    ws['B2'] = len(sorted_tests)
    ws['B3'] = failed_tests
    ws['B4'] = error_tests
    ws['B5'] = skipped_tests
    ws['B6'] = passed_tests

    # Set DCI Job-ID
    ws['A7'] = 'Job-Id'
    ws['B7'] = 'https://www.distributed-ci.io/jobs/' + dci_jobid
    ws['A8'] = ''
    ws['A9'] = ''

    # Set summary styles
    summary_fill = {
        'Summary': styles['blue_fill'],
        'Total': PatternFill(start_color='00FFFF', end_color='00FFFF', fill_type='solid'),
        'Failed': styles['red_fill'],
        'Error': styles['dark_red_fill'],
        'Skipped': styles['orange_fill'],
        'Passed': styles['green_fill'],
        'Job-Id': styles['yellow_fill']
    }
    
    for row in ws.iter_rows(min_row=1, max_row=8):
        for cell in row:
            cell.font = styles['bold_font']
            if isinstance(cell.value, str) and cell.value in summary_fill:
                cell.fill = summary_fill[cell.value]

def add_version_information(ws: Worksheet, data: Dict[str, Any], 
                           styles: Dict[str, Any]) -> None:
    """Add version information to the worksheet."""
    try:
        # Extract the values for k8s, ocClient, ocp, and cert with safe access
        versions = data.get('claim', {}).get('versions', {})
        
        k8s_value = versions.get('k8s', 'N/A')
        oc_client_value = versions.get('ocClient', 'N/A')
        ocp_value = versions.get('ocp', 'N/A')
        cert_value = versions.get('certSuite', 'N/A')
        claim_format = versions.get('claimFormat', 'N/A')
        cert_git_commit = versions.get('certSuiteGitCommit', 'N/A')

        ws['C1'] = 'Component'
        ws['D1'] = 'Version'
        ws['C2'] = 'K8S'
        ws['C3'] = 'ocClient'
        ws['C4'] = 'OCP'
        ws['C5'] = 'CERTSUIT'
        ws['C6'] = 'claimF'
        ws['C7'] = 'certGitCom'

        ws['D2'] = k8s_value
        ws['D3'] = oc_client_value
        ws['D4'] = ocp_value
        ws['D5'] = cert_value
        ws['D6'] = claim_format
        ws['D7'] = cert_git_commit
    except Exception as e:
        print(f"Warning: Error adding version information: {e}")
        # Continue with empty values if version info is not available

def apply_final_formatting(ws: Worksheet, styles: Dict[str, Any]) -> None:
    """Apply final formatting and styling to the worksheet."""
    # Center alignment for specific rows
    rows_to_center = [1, 9]
    for row_number in rows_to_center:
        for cell in ws[row_number]:
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)

    # Define border style
    border = Border(
        left=Side(border_style='thin', color='D3D3D3'),
        right=Side(border_style='thin', color='D3D3D3'),
        top=Side(border_style='thin', color='D3D3D3'),
        bottom=Side(border_style='thin', color='D3D3D3')
    )

    # Set version styles for version info section
    for i, cell in enumerate(ws['C']):
        if i in [1, 2, 3, 4, 5, 6]:
           cell.border = border
           cell.fill = styles['yellow_fill']

    for i, cell in enumerate(ws['A']):
        if i in [1, 2, 3, 4, 5, 6]:
           cell.border = border

    # Set border to State Column C
    for i in range(10, ws.max_row + 1):
        cell = ws[f'C{i}']
        cell.border = border

    # Directly set the fill for the first cell in columns 'C' and 'D'
    ws['C1'].fill = styles['light_green_fill']
    ws['D1'].fill = styles['light_green_fill']

    # Set alignment for specific cells
    for cell in ws['B'][ws.min_row-1:ws.max_row-6]:
        cell.alignment = Alignment(horizontal='left', vertical='center')

    for cell in ws['D'][ws.min_row-2:ws.max_row-5]:
        cell.alignment = Alignment(horizontal='left', vertical='center')

    # Set column widths
    ws.column_dimensions['B'].width = 100
    ws.column_dimensions['D'].width = 100
    ws.column_dimensions['F'].width = 100
    ws.column_dimensions['G'].width = 100
    ws.column_dimensions['H'].width = 100
    ws.column_dimensions['C'].width = 11

    # Set special font for Category Classification
    blue_bold_font = Font(name='Arial', bold=False, color="0000FF")
    keywords = ["Extended:", "FarEdge:", "NonTelco:", "Telco:"]
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=6):
        for cell in row:
            # Check if the cell contains any of the specified strings
            if isinstance(cell.value, str) and any(keyword in cell.value for keyword in keywords):
                cell.font = blue_bold_font

    # Set row heights
    for row_num in range(10, 200): 
        ws.row_dimensions[row_num].height = 30

def generate_cert_test_excel_report(input_claim: str, output_file: str, dci_jobid: str):
    """Generate Excel report from CertSuite claim JSON file."""
    try:
        # Load JSON data from file
        if not os.path.exists(input_claim):
            raise FileNotFoundError(f"Input file not found: {input_claim}")
            
        with open(input_claim, 'r', encoding='utf-8') as f:
            data = json.load(f)
            
        # Validate required data structure
        if 'claim' not in data:
            raise ValueError("Invalid claim file: missing 'claim' section")
            
        if 'results' not in data.get('claim', {}):
            raise ValueError("Invalid claim file: missing 'results' section")

        # Extract test results and process them
        sorted_tests, total_failed, total_error, total_skipped, total_passed = extract_test_results(data)

        # Create workbook and worksheet
        wb, ws = create_workbook_and_worksheet(output_file)

        # Add test results to worksheet
        add_test_results_to_worksheet(ws, sorted_tests)

        # Apply basic styling
        styles = apply_basic_styling(ws)

        # Set column formatting
        set_column_formatting(ws)

        # Add summary section
        add_summary_section(ws, sorted_tests, total_failed, total_error, total_skipped, total_passed, dci_jobid, styles)

        # Add version information
        add_version_information(ws, data, styles)

        # Apply final formatting
        apply_final_formatting(ws, styles)

        # Save workbook to file
        wb.save(output_file)
        
    except FileNotFoundError as e:
        print(f"Error: {e}")
        sys.exit(1)
    except json.JSONDecodeError as e:
        print(f"Error: Invalid JSON in file {input_claim}: {e}")
        sys.exit(1)
    except ValueError as e:
        print(f"Error: {e}")
        sys.exit(1)
    except Exception as e:
        print(f"Unexpected error generating Excel report: {e}")
        sys.exit(1)


def read_dcirc_env_variables(dcirc_path: str = "dcirc.sh") -> Dict[str, str]:
    """Read environment variables from dcirc.sh file."""
    env_vars = {}
    
    if not os.path.exists(dcirc_path):
        raise FileNotFoundError(f"DCI configuration file not found: {dcirc_path}")
    
    try:
        with open(dcirc_path, 'r') as f:
            content = f.read()
            
        # Parse export statements using regex
        # Pattern to match: export VAR_NAME="value" or export VAR_NAME=value
        pattern = r'export\s+([A-Z_]+)=(?:"([^"]*)"|([^\s]+))'
        
        for match in re.finditer(pattern, content):
            var_name = match.group(1)
            # Use quoted value if present, otherwise use unquoted value
            var_value = match.group(2) if match.group(2) is not None else match.group(3)
            env_vars[var_name] = var_value
            
        return env_vars
        
    except Exception as e:
        raise ValueError(f"Error reading DCI configuration file {dcirc_path}: {e}")

def get_dci_environment() -> Dict[str, str]:
    """Get DCI environment variables from environment or dcirc.sh file."""
    required_vars = ['DCI_CLIENT_ID', 'DCI_API_SECRET', 'DCI_CS_URL']
    env_vars = {}
    
    # First, try to get from environment variables
    missing_vars = []
    for var in required_vars:
        value = os.environ.get(var)
        if value:
            env_vars[var] = value
        else:
            missing_vars.append(var)
    
    # If some variables are missing, try to read from dcirc.sh
    if missing_vars:
        print(f"Missing environment variables: {', '.join(missing_vars)}")
        print("Attempting to read from dcirc.sh file...")
        
        try:
            dcirc_vars = read_dcirc_env_variables()
            
            # Update with values from dcirc.sh
            for var in missing_vars:
                if var in dcirc_vars:
                    env_vars[var] = dcirc_vars[var]
                    print(f"✓ Found {var} in dcirc.sh")
                else:
                    raise ValueError(f"Required DCI variable '{var}' not found in dcirc.sh")
                    
        except Exception as e:
            print(f"Error reading dcirc.sh: {e}")
            raise ValueError(f"Cannot find required DCI environment variables. "
                           f"Please either:\n"
                           f"1. Export the variables: {', '.join(missing_vars)}\n"
                           f"2. Ensure dcirc.sh exists with proper export statements")
    
    return env_vars

def download_dci_cert_claim_json(input_claim: str, job_id: str) -> None:
    """Download claim.json file from DCI control server using job ID."""
    try:
        # Get DCI environment variables (from environment or dcirc.sh)
        env_vars = get_dci_environment()
        
        # Create environment dictionary for subprocess
        env = os.environ.copy()  # Start with current environment
        env.update(env_vars)     # Add DCI variables
        
        print(f"Using DCI server: {env_vars.get('DCI_CS_URL', 'N/A')}")
        
        # run the dcictl command and capture the output
        cmd = ['dcictl', 'file-list', job_id, '--limit', '200']
        process = subprocess.Popen(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE, env=env)
        stdout, stderr = process.communicate()
        
        if process.returncode != 0:
            print(f"Error running dcictl file-list: {stderr.decode('utf-8')}")
            sys.exit(1)

        # extract the file IDs from the output using grep
        output_lines = stdout.decode('utf-8').split('\n')
        claim_file_lines = [line for line in output_lines if 'claim.json' in line and 'text/plain' in line]
        
        if not claim_file_lines:
            print("No claim.json files found in the job")
            sys.exit(1)
            
        file_ids = [line.split()[1] for line in claim_file_lines]
        print(f"Found {len(file_ids)} claim.json file(s): {file_ids}")

        # run the dcictl command to download the claim.json file(s)
        cmd = ['dcictl', 'job-download-file', job_id, '--file-id', ','.join(file_ids), '--target', input_claim]
        result = subprocess.run(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE, env=env)

        # check if the command was successful
        if result.returncode == 0:
            print("dcictl command succeeded to download claim.json")
        else:
            print(f"dcictl command failed to download claim.json with return code {result.returncode}")
            print(f"Error: {result.stderr.decode('utf-8')}")
            sys.exit(1)
            
    except Exception as e:
        print(f"Error downloading claim.json: {e}")
        sys.exit(1)

def check_file_exists(filename: str) -> None:
    """Check if a file exists and exit if it doesn't."""
    if os.path.isfile(filename):
        print(f"{filename} exists")
    else:
        print(f"{filename} does not exist, this file requires to access DCI Control Server!")
        sys.exit(0)

def check_tool_not_exists(tool_name: str) -> bool:
    """Check if a command-line tool exists."""
    try:
        subprocess.check_output([tool_name, "--version"], stderr=subprocess.STDOUT)
        return False
    except FileNotFoundError:
        return True
    except subprocess.CalledProcessError:
        return True

if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Parse CertSuite claim.json and save as Excel workbook')
    parser.add_argument('-i', '--input-file', type=str, required=True, 
                       help='Provide the input to claim.json file')
    parser.add_argument('-j', '--job-id', type=str, 
                       help='Provide the DCI job-id if download automatic, otherwise do not specify to parse offline by copy claim.json here')
    parser.add_argument('-o', '--output-file', type=str, required=True, 
                       help='Path to output Excel file e.g. result.xlsx')

    args = parser.parse_args()

    # Check if dcirc.sh exists
    check_file_exists("dcirc.sh")

    # Check if dcictl is installed
    tool_name = "dcictl"
    if check_tool_not_exists(tool_name):
        print(f"{tool_name} is not installed.")
        # sys.exit(0)  # Commented out to allow offline mode

    # Download claim.json from DCI control server using job-id if provided
    # Otherwise, user can use claim.json as offline parser
    if args.job_id is not None:
        download_dci_cert_claim_json(args.input_file, args.job_id)
    else:
        args.job_id = "12345"  # Default job ID for offline mode
        print("You are using claim.json as offline parser!")

    # Generate the Excel report
    generate_cert_test_excel_report(args.input_file, args.output_file, args.job_id)
    print(f"Successfully parsed claim.json and generated Excel report: {args.output_file}")

